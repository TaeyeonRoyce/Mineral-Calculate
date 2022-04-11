import pandas as pd
import openpyxl
import mineralCroller
import formulaHandler
import json
import numpy as np

databaseSource = "mineralDB.xlsx"
elementsDataFrame = pd.read_excel(
    databaseSource, sheet_name="Elements", index_col="Symbol"
)

mineralDataFrame = pd.read_excel(
    databaseSource, sheet_name="DB_1", header=3, index_col="광물 이름"
)

massDataFrame = elementsDataFrame["Mass per mole"]
metalDataFrame = elementsDataFrame["비금속"]
priceDataFrame = elementsDataFrame["Price (USD / kg)"]
gsDataFrame = mineralDataFrame["비중"]


def getMineralNameList():
    # DB정보 명세
    sheetName = "DB_1"
    startRow = 3
    column = "B"

    # Excel DB에서 광물 이름 추출(in English)
    mineralNameDataFrame = pd.read_excel(
        databaseSource, sheet_name=sheetName, header=startRow, usecols=column
    )
    mineralNameList = mineralNameDataFrame.loc[0:80, "광물 이름"].tolist()
    return mineralNameList


def getElementMassPerMole(element):
    mass = massDataFrame.loc[[element]]
    return float(mass)


def getFormulaList():
    # DB정보 명세
    sheetName = "Result"
    startRow = 3
    column = "D"

    # Excel DB에서 화학식 추출
    mineralNameDataFrame = pd.read_excel(
        databaseSource, sheet_name=sheetName, header=startRow, usecols=column
    )
    formulaList = mineralNameDataFrame.loc[0:99, "화학식"].tolist()
    return formulaList


def getElementPricePerKG(element):
    price = priceDataFrame.loc[[element]]
    return float(price)


def getMineralGs(mineral):
    gs = str(gsDataFrame.loc[[mineral]][0])
    gs = gs.split(" ~ ")
    if len(gs) > 1:
        return round(((float(gs[0]) + float(gs[1])) / 2), 2)
    return float(gs[0])


def isMetal(element):
    value = metalDataFrame.loc[[element]]
    if float(value) == 1.0:
        return False
    return True


def saveChemicalFormula():
    wb = openpyxl.load_workbook(databaseSource)
    sheet = wb.active
    mineralList = getMineralNameList()
    column = "D"
    for i in range(len(mineralList)):
        row = i + 5
        cell = column + str(row)
        if sheet[cell].value == None:
            sheet[cell].value = mineralCroller.getChemicalFormulaBy(mineralList[i])
            print(sheet[cell].value)
            wb.save(databaseSource)


def saveComponentsDict():
    wb = openpyxl.load_workbook(databaseSource)
    sheet = wb.active
    formulaList = getFormulaList()
    column = "F"
    for i in range(len(formulaList)):
        if formulaList[i] == "Null":
            continue
        row = i + 5
        cell = column + str(row)
        if sheet[cell].value == None:
            sheet[cell].value = json.dumps(
                formulaHandler.findElemetnsFromFormula(formulaList[i])
            )
            print("{0} : {1}".format(formulaList[i], sheet[cell].value))
            wb.save(databaseSource)


def saveComponentsDictMetal():
    wb = openpyxl.load_workbook(databaseSource)
    sheet = wb.active
    formulaList = getFormulaList()
    column = "H"
    for i in range(len(formulaList)):
        if formulaList[i] == "Null":
            continue
        row = i + 5
        cell = column + str(row)
        if sheet[cell].value == None:
            componentsDict = formulaHandler.findElemetnsFromFormula(formulaList[i])
            sheet[cell].value = json.dumps(filterNoneMetal(componentsDict))
            print("save : ", sheet[cell].value)
            wb.save(databaseSource)


def filterNoneMetal(componentsDict):
    metalDict = {}
    for element in componentsDict:
        if isMetal(element):
            metalDict[element] = componentsDict[element]
    return metalDict


def savePricePerKG():
    wb = openpyxl.load_workbook(databaseSource)
    sheet = wb.active
    formulaList = getFormulaList()
    column = "J"
    for i in range(len(formulaList)):
        if formulaList[i] == "Null":
            continue
        row = i + 5
        cell = column + str(row)
        if sheet[cell].value == None:
            componentsDict = formulaHandler.findElemetnsFromFormula(formulaList[i])
            metalDict = filterNoneMetal(componentsDict)
            pricePerKg = round(formulaHandler.calPriceElementsDict(metalDict), 2)
            sheet[cell].value = json.dumps(pricePerKg)
            print("save : ", sheet[cell].value)
            wb.save(databaseSource)


def printPriceListDESC():
    # DB정보 명세
    sheetName = "Result"
    startRow = 3
    mineralNameIndex = 1
    PricePerKGIndex = 9

    # Excel DB에서 화학식 추출
    priceDataFrame = pd.read_excel(
        databaseSource,
        sheet_name=sheetName,
        header=startRow,
        usecols=[mineralNameIndex, PricePerKGIndex],
    )
    print(priceDataFrame.sort_values(by=["단위 질량 당 가격(USD/kg)"], ascending=False))


def saveGs():
    wb = openpyxl.load_workbook(databaseSource)
    sheet = wb.active
    mineralList = getMineralNameList()
    column = "F"
    for i in range(len(mineralList)):
        row = i + 5
        cell = column + str(row)
        if sheet[cell].value == None:
            print(mineralList[i])
            sheet[cell].value = mineralCroller.getGSBy(mineralList[i])
            print(sheet[cell].value)
            wb.save(databaseSource)


def savePricePerVolume():
    wb = openpyxl.load_workbook(databaseSource)
    sheet = wb.active

    minPriceMassDataFrame = pd.read_excel(
        databaseSource,
        sheet_name="Result",
        header=3,
        usecols=["광물 이름", "단위 질량 당 가격(USD/kg)"],
    )
    column = "L"
    i = 0
    for idx in minPriceMassDataFrame.index:
        if minPriceMassDataFrame.loc[idx, "단위 질량 당 가격(USD/kg)"] == "Null":
            i += 1
            continue
        row = i + 5
        cell = column + str(row)
        if sheet[cell].value == None:
            gs = getMineralGs(minPriceMassDataFrame.loc[idx, "광물 이름"])
            pricePerVolume = round(
                gs * minPriceMassDataFrame.loc[idx, "단위 질량 당 가격(USD/kg)"], 2
            )
            sheet[cell].value = pricePerVolume
            print("save : ", sheet[cell].value)
            wb.save(databaseSource)
        i += 1


def printPriceVolumeDESC():
    # DB정보 명세
    sheetName = "Result"
    startRow = 3
    mineralNameIndex = 1
    PricePerVolumeIndex = 11

    # Excel DB에서 화학식 추출
    priceDataFrame = pd.read_excel(
        databaseSource,
        sheet_name=sheetName,
        header=startRow,
        usecols=[mineralNameIndex, PricePerVolumeIndex],
    )
    descResult = priceDataFrame.sort_values(by=["단위 부피 당 가격(USD/cm3)"], ascending=False)

    wb = openpyxl.load_workbook(databaseSource)
    sheet = wb.active
    i = 5
    columnMin = "B"
    columnPrice = "D"
    for idx in descResult.index:
        cellMin = columnMin + str(i)
        cellPrice = columnPrice + str(i)
        sheet[cellMin].value = descResult.loc[idx, "광물 이름"]
        sheet[cellPrice].value = descResult.loc[idx, "단위 부피 당 가격(USD/cm3)"]
        print(sheet[cellMin].value)
        print(sheet[cellPrice].value)
        wb.save(databaseSource)
        i += 1


printPriceVolumeDESC()
